"""매출내역서 엑셀 내보내기 - 실무 양식 호환"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.calculations import ENGINE_OIL_UNIT_PRICE, fmt_money


def export_sales_excel(rows: list[dict], year: int, month: int) -> bytes:
    """매출내역서 엑셀 생성 → bytes 반환"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}월 매출내역서"

    # ── 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── 제목
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"자동차 점검·정비 매출내역서 ({year}년 {month}월)"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = center

    # ── 엔진오일 단가 안내
    ws.merge_cells("A2:R2")
    ws["A2"].value = f"엔진오일 실단가: {ENGINE_OIL_UNIT_PRICE:,}원/리터"
    ws["A2"].font = Font(size=9, color="666666")

    # ── 컬럼 헤더 (실무 양식 일치)
    headers = [
        "순번", "입고일", "출고일", "입고분류", "차량번호", "차량모델",
        "부품금액\n(엔진오일 제외)", "견인비", "보험료",
        "엔진오일\n(리터)", "엔진오일\n(단가)", "엔진오일\n(금액)",
        "총부품금액\n(보험료·견인비 포함)", "부품수익", "기술료",
        "부가가치세", "총  계", "도장금액"
    ]
    col_widths = [6, 12, 12, 14, 12, 16,
                  14, 10, 10,
                  10, 10, 12,
                  16, 10, 10,
                  10, 14, 10]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[3].height = 36

    # ── 데이터
    money_cols = {7, 8, 9, 12, 13, 14, 15, 16, 17, 18}
    num_fmt = '#,##0'

    for ri, row in enumerate(rows, start=4):
        vals = [
            ri - 3,
            row.get("intake_date"),
            row.get("actual_out"),
            row.get("intake_type"),
            row.get("plate_number"),
            row.get("model"),
            row.get("parts_amount", 0) or 0,
            row.get("towing_fee", 0) or 0,
            row.get("insurance_fee", 0) or 0,
            row.get("engine_oil_liter", 0) or 0,
            row.get("engine_oil_unit", ENGINE_OIL_UNIT_PRICE) or ENGINE_OIL_UNIT_PRICE,
            row.get("engine_oil_amount", 0) or 0,
            row.get("total_parts", 0) or 0,
            (row.get("total_parts", 0) or 0) - (row.get("parts_amount", 0) or 0),  # 부품수익 추정
            row.get("tech_fee", 0) or 0,
            row.get("vat_amount", 0) or 0,
            row.get("total_amount", 0) or 0,
            row.get("paint_amount", 0) or 0,
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = center
            if ci in money_cols and val is not None:
                cell.number_format = num_fmt

    # ── 합계 행
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="합 계").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = center
    ws.merge_cells(f"A{total_row}:F{total_row}")
    for ci in range(7, 19):
        col_letter = get_column_letter(ci)
        ws.cell(
            row=total_row, column=ci,
            value=f"=SUM({col_letter}4:{col_letter}{total_row-1})"
        ).number_format = num_fmt
        ws.cell(row=total_row, column=ci).font = Font(bold=True)
        ws.cell(row=total_row, column=ci).alignment = center
        ws.cell(row=total_row, column=ci).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_work_list_excel(rows: list[dict]) -> bytes:
    """작업진행내역 엑셀 내보내기"""
    df = pd.DataFrame(rows)
    col_map = {
        "plate_number": "차량번호", "model": "차량모델",
        "intake_date": "입고일", "expected_out": "출고예정일",
        "actual_out": "실제출고일", "intake_type": "입고분류",
        "status": "작업상태", "customer_name": "고객명",
        "customer_phone": "연락처", "aos_claimed": "AOS",
        "insurance_paid": "보험입금", "memo": "비고",
    }
    df = df.rename(columns=col_map)
    available = [c for c in col_map.values() if c in df.columns]
    df = df[available]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="작업진행내역")
    buf.seek(0)
    return buf.read()
